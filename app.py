# ============================================================
#  JBS SNIPER — app.py  v7 DEFINITIVO
#  3 parsers: icontemplados_detalhe | icontemplados_cards | generico
# ============================================================

import streamlit as st
import pandas as pd
import re
import itertools
import hashlib
import os
from io import BytesIO
from fpdf import FPDF
from datetime import datetime

_favicon = "logo_pdf.png" if os.path.exists("logo_pdf.png") else "🎯"

st.set_page_config(
    page_title="JBS SNIPER",
    page_icon=_favicon,
    layout="wide",
    initial_sidebar_state="collapsed",
)

GOLD   = "#84754e"
BEIGE  = "#ecece4"
BG     = "#0e1117"
CARD   = "#1c1f26"
BORDER = "#2e3340"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow:wght@300;400;600;700&family=Barlow+Condensed:wght@400;600;700&display=swap');
html, body, [class*="css"] {{ font-family: 'Barlow', sans-serif; }}
.stApp {{ background-color: {BG}; color: {BEIGE}; }}
.stButton > button {{
    width: 100%;
    background: linear-gradient(135deg, {GOLD} 0%, #6b5e3d 100%);
    color: #fff; border: none; border-radius: 8px;
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 1.05rem; font-weight: 700;
    letter-spacing: 2px; text-transform: uppercase;
    padding: 14px 0; transition: all .2s ease;
    box-shadow: 0 4px 15px rgba(132,117,78,.3);
}}
.stButton > button:hover {{ transform: translateY(-2px); box-shadow: 0 6px 20px rgba(132,117,78,.45); }}
.stButton > button:active {{ transform: translateY(0); }}
.stTextArea textarea, .stNumberInput input, .stSelectbox > div > div {{
    background-color: {CARD} !important; color: {BEIGE} !important;
    border: 1px solid {BORDER} !important; border-radius: 8px !important;
    font-family: 'Barlow', sans-serif !important;
}}
.stTextArea textarea:focus, .stNumberInput input:focus {{
    border-color: {GOLD} !important;
    box-shadow: 0 0 0 2px rgba(132,117,78,.25) !important;
}}
label, .stSelectbox label, .stNumberInput label, .stTextArea label {{
    color: {BEIGE} !important; font-size: .78rem !important;
    font-weight: 600 !important; letter-spacing: 1px !important;
    text-transform: uppercase !important;
}}
h1, h2, h3, h4 {{
    font-family: 'Barlow Condensed', sans-serif !important;
    color: {GOLD} !important; letter-spacing: 1px;
}}
.stDataFrame {{ border: 1px solid {BORDER}; border-radius: 8px; overflow: hidden; }}
.streamlit-expanderHeader {{
    background-color: {CARD} !important; border: 1px solid {BORDER} !important;
    border-radius: 8px !important; color: {BEIGE} !important;
    font-family: 'Barlow Condensed', sans-serif !important;
    font-weight: 600 !important; letter-spacing: 1px !important;
}}
[data-testid="metric-container"] {{
    background-color: {CARD}; border: 1px solid {BORDER};
    border-radius: 10px; padding: 14px 18px;
}}
[data-testid="metric-container"] label {{
    color: {GOLD} !important; font-size: .72rem !important; letter-spacing: 1.5px !important;
}}
[data-testid="metric-container"] [data-testid="stMetricValue"] {{
    color: {BEIGE} !important;
    font-family: 'Barlow Condensed', sans-serif !important; font-size: 1.6rem !important;
}}
hr {{ border-color: {BORDER}; margin: 8px 0 20px; }}
.stProgress > div > div > div {{ background-color: {GOLD} !important; }}
::-webkit-scrollbar {{ width: 6px; height: 6px; }}
::-webkit-scrollbar-track {{ background: {BG}; }}
::-webkit-scrollbar-thumb {{ background: {BORDER}; border-radius: 3px; }}
::-webkit-scrollbar-thumb:hover {{ background: {GOLD}; }}
</style>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════
#  CABEÇALHO
# ════════════════════════════════════════════════════════════
col_logo, col_title = st.columns([1, 5])
with col_logo:
    if os.path.exists("logo_app.png"):
        st.image("logo_app.png", width=200)
    else:
        st.markdown(f"<h1 style='font-size:3rem;margin:0;padding-top:8px'>{_favicon}</h1>", unsafe_allow_html=True)
with col_title:
    st.markdown(f"""
    <div style='padding-top:10px'>
        <h1 style='margin:0;font-size:2.4rem;letter-spacing:4px'>SISTEMA SNIPER</h1>
        <p style='color:{BEIGE};opacity:.6;margin:0;font-size:.9rem;letter-spacing:2px;text-transform:uppercase'>
            JBS Contempladas · Inteligência Comercial
        </p>
    </div>
    """, unsafe_allow_html=True)
st.markdown(f"<hr style='border:1px solid {GOLD};margin-top:4px'>", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
#  FORMATAÇÃO BR
# ════════════════════════════════════════════════════════════
def fmt_brl(valor) -> str:
    try:
        s = f"{float(valor):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        return f"R$ {s}"
    except Exception:
        return "R$ 0,00"

def fmt_pct(valor) -> str:
    try:
        return f"{float(valor):.2f}%".replace(".",",")
    except Exception:
        return "0,00%"

def fmt_pct_curto(valor) -> str:
    try:
        return f"{float(valor):.0f}%"
    except Exception:
        return "0%"


# ════════════════════════════════════════════════════════════
#  UTILITÁRIOS
# ════════════════════════════════════════════════════════════
def limpar_moeda(texto) -> float:
    if not texto: return 0.0
    try:
        t = (str(texto).lower().strip()
             .replace('\xa0','').replace('&nbsp;','')
             .replace('r$','').replace(' ',''))
        t = re.sub(r'[^\d\.,]', '', t)
        if not t: return 0.0
        if ',' in t and '.' in t:
            return float(t.replace('.','').replace(',','.'))
        if ',' in t:
            p = t.split(',')
            return float(t.replace(',','.')) if len(p)==2 and len(p[1])<=2 else float(t.replace(',',''))
        if '.' in t:
            p = t.split('.')
            return float(t) if len(p)==2 and len(p[1])==2 else float(t.replace('.',''))
        return float(t)
    except Exception:
        return 0.0

def _detectar_tipo(b: str) -> str:
    if any(k in b for k in ('imóvel','imovel','apartamento','casa','terreno','comercial')): return "Imóvel"
    if any(k in b for k in ('caminhão','caminhao','pesado','truck','ônibus','onibus')): return "Pesados"
    if any(k in b for k in ('automóvel','automovel','veículo','veiculo','carro','moto')): return "Automóvel"
    return "Geral"


# ════════════════════════════════════════════════════════════
#  LISTAS GLOBAIS
# ════════════════════════════════════════════════════════════
_ADMINS = [
    'ITAÚ AUTO','ITAU AUTO','BRADESCO AUTO','BRADESCO IMÓVEIS',
    'BRADESCO','SANTANDER','ITAÚ','ITAU','PORTO SEGURO','PORTO',
    'CAIXA','BANCO DO BRASIL','BB','RODOBENS','EMBRACON',
    'ANCORA','ÂNCORA','MYCON','SICREDI','SICOOB','MAPFRE',
    'HS','YAMAHA','ZEMA','BANCORBRÁS','BANCORBRAS','SERVOPA',
    'WOOP','SOMPO','MAGALU',
]

_BANCOS_IC = [
    'Itaú Auto','Itau Auto','Bradesco Auto','Bradesco Imóveis',
    'Bradesco','Santander','Porto Seguro','Caixa',
    'Banco do Brasil','BB','Rodobens','Embracon','Mycon',
    'Sicredi','Sicoob','Mapfre','Yamaha','Zema',
    'Magalu','Woop','Sompo','Ancora',
]
_BANCOS_IC_LOWER = {b.lower() for b in _BANCOS_IC}

_IGNORAR_IC = {
    'selecionar','detalhes',
    'código:','codigo:','directions_car','directions_home','directions_bus',
    'warning','info','check_circle','error',
    'baixar pdf','baixar excel','selecionar todas','criar filtro',
    'compartilhar','somar','início','inicio','todos',
    'automóveis','automoveis','itaú consórcio','itau consorcio',
    'cartas contempladas para veículo','cartas contempladas para veiculo',
    'cartas contempladas para imovel','cartas contempladas para imóvel',
    'imóvel','imovel','veículo','veiculo',
    # NOTA: 'reservada','negociar','disponível' NAO estão aqui —
    # precisam chegar ao parser para detectar status da cota
}

_RE_CREDITO = re.compile(r'(?:cr[eé]dito|bem|valor[\s_]do[\s_]bem|valor[\s_]carta)[^\d\n]{0,25}?R\$\s*([\d\.,]+)', re.IGNORECASE)
_RE_ENTRADA = re.compile(r'(?:entrada|[aá]gio|lance[\s_]fixo|pago|quero)[^\d\n]{0,25}?R\$\s*([\d\.,]+)', re.IGNORECASE)
_RE_PARCELA = re.compile(r'(\d+)\s*[xX]\s*R?\$?\s*([\d\.,]+)', re.IGNORECASE)
_RE_MOEDA   = re.compile(r'R\$\s*([\d\.,]+)', re.IGNORECASE)


# ════════════════════════════════════════════════════════════
#  DETECTOR DE FORMATO
# ════════════════════════════════════════════════════════════
def _detectar_formato(texto: str) -> str:
    # 1. iContemplados com Detalhes expandidos
    if re.search(r'saldo\s+devedor\s*:', texto, re.IGNORECASE):
        return "icontemplados_detalhe"
    # 2. iContemplados cards (sem detalhe): tem banco + "Entrada:\n" + "Parcelas:\n"
    tem_entrada  = bool(re.search(r'^entrada\s*:?\s*$', texto, re.IGNORECASE | re.MULTILINE))
    tem_parcelas = bool(re.search(r'^parcelas?\s*:?\s*$', texto, re.IGNORECASE | re.MULTILINE))
    tem_banco    = any(b in texto for b in _BANCOS_IC)
    if tem_banco and tem_entrada and tem_parcelas:
        return "icontemplados_cards"
    # 3. Genérico
    return "generico"


# ════════════════════════════════════════════════════════════
#  PARSER 1 — iContemplados com Detalhes expandidos
# ════════════════════════════════════════════════════════════
def _extrair_icontemplados_detalhe(texto: str, tipo_sel: str) -> list:
    lista, id_c = [], 1
    blocos = re.split(r'(?i)(?=administradora\s*:)', texto)
    for bloco in blocos:
        if 'administradora' not in bloco.lower(): continue
        if len(bloco.strip()) < 30: continue
        try:
            m = re.search(r'administradora\s*:\s*\*?\*?([^\n\*]+)', bloco, re.IGNORECASE)
            admin_raw = m.group(1).strip() if m else "OUTROS"
            admin = admin_raw.upper()
            for adm in _ADMINS:
                if adm.lower() in admin_raw.lower():
                    admin = adm.upper(); break

            m = re.search(r'cr[eé]dito\s*:\s*\*?\*?\s*R\$\s*([\d\.,]+)', bloco, re.IGNORECASE)
            credito = limpar_moeda(m.group(1)) if m else 0.0
            if credito <= 0: continue

            m = re.search(r'segmento\s*:\s*\*?\*?([^\n\*]+)', bloco, re.IGNORECASE)
            tipo_raw = m.group(1).strip().lower() if m else ""
            if any(k in tipo_raw for k in ('imóvel','imovel','imov')): tipo = "Imóvel"
            elif any(k in tipo_raw for k in ('pesado','caminhão','caminhao')): tipo = "Pesados"
            elif any(k in tipo_raw for k in ('veículo','veiculo','auto','carro','moto')): tipo = "Automóvel"
            else: tipo = "Imóvel" if credito > 250000 else "Automóvel"

            if tipo_sel not in ("Todos","Geral") and tipo != tipo_sel: continue

            m = re.search(r'entrada\s*:\s*\*?\*?\s*R\$\s*([\d\.,]+)', bloco, re.IGNORECASE)
            entrada = limpar_moeda(m.group(1)) if m else 0.0
            if entrada <= 0: continue

            m = re.search(r'saldo\s+devedor\s*:\s*\*?\*?\s*R\$\s*([\d\.,]+)', bloco, re.IGNORECASE)
            saldo = limpar_moeda(m.group(1)) if m else 0.0

            m = re.search(r'parcelas?\s*:\s*\*?\*?\s*(\d+)\s*[xX]\s*R?\$?\s*([\d\.,]+)', bloco, re.IGNORECASE)
            parcela    = limpar_moeda(m.group(2)) if m else 0.0
            n_parcelas = int(m.group(1)) if m else 0

            if saldo <= 0:
                saldo = max(credito * 1.25 - entrada, credito * 0.20)

            lista.append({
                'ID': id_c, 'Admin': admin, 'Tipo': tipo,
                'Crédito': credito, 'Entrada': entrada,
                'Parcela': parcela, 'NParcelas': n_parcelas,
                'Saldo': saldo, 'CustoTotal': entrada + saldo,
                'EntradaPct': entrada / credito,
            })
            id_c += 1
        except Exception:
            continue
    return lista


# ════════════════════════════════════════════════════════════
#  PARSER 2 — iContemplados cards (Ctrl+A sem expandir Detalhes)
#
#  Estrutura real do Ctrl+A:
#    directions_car        ← ícone (ignora)
#    Itaú Auto             ← ADMIN (linha = nome de banco)
#    R$ 23.050,04          ← CRÉDITO (linha seguinte começa com R$)
#    Entrada:              ← rótulo sozinho
#    R$ 5.200,00           ← ENTRADA (linha seguinte)
#    Parcelas:             ← rótulo sozinho
#    28 x R$ 913,68        ← PARCELAS (linha seguinte)
#    Código:               ← ignora
#    593                   ← ignora (número curto)
#    Selecionar            ← ignora
# ════════════════════════════════════════════════════════════
def _extrair_icontemplados_cards(texto: str, tipo_sel: str) -> list:
    """Parser iContemplados cards.
    Passo 1: extrai codigos reservados do texto bruto.
    Passo 2: extrai cotas e descarta as reservadas pelo codigo.
    """
    lista, id_c = [], 1

    # Passo 1: codigos reservados — padrao: [numero]\nSelecionar\nDetalhes\nReservada
    codigos_reservados = set(re.findall(
        r'(\d+)\s*\nSelecionar\s*\nDetalhes\s*\nReservada',
        texto, re.IGNORECASE
    ))
    # Variacao: Codigo:\n[numero]\n...Reservada
    for m in re.finditer(
        r'[Cc][oó]digo:\s*\n(\d+)\s*\n.*?Reservada',
        texto, re.DOTALL
    ):
        codigos_reservados.add(m.group(1))

    # Passo 2: tipo default pelo cabecalho
    header = texto[:500].lower()
    if any(k in header for k in ('para veículo','para veiculo','veículo','veiculo')):
        tipo_default = "Automóvel"
    elif any(k in header for k in ('para imóvel','para imovel','imóvel','imovel')):
        tipo_default = "Imóvel"
    else:
        tipo_default = "Automóvel"

    # Passo 3: limpa linhas — remove UI mas mantem numeros (codigos das cotas)
    _SKIP = {
        'selecionar','detalhes','reservada','negociar',
        'disponível','disponivel','código:','codigo:',
        'directions_car','directions_home','directions_bus',
        'warning','info','check_circle','error',
        'baixar pdf','baixar excel','selecionar todas','criar filtro',
        'compartilhar','somar','início','inicio','todos',
        'automóveis','automoveis','itaú consórcio','itau consorcio',
        'cartas contempladas para veículo','cartas contempladas para veiculo',
        'cartas contempladas para imovel','cartas contempladas para imóvel',
        'imóvel','imovel','veículo','veiculo',
    }
    linhas = []
    for ln in texto.replace('\r', '').split('\n'):
        ln = ln.strip()
        if not ln: continue
        if ln.lower() in _SKIP: continue
        linhas.append(ln)

    # Passo 4: extrai cotas
    i = 0
    while i < len(linhas):
        ln = linhas[i]

        if (ln.lower() in _BANCOS_IC_LOWER
                and i + 1 < len(linhas)
                and re.match(r'^R\$\s*[\d\.]+,\d{2}$', linhas[i + 1])):

            admin_raw  = ln
            credito    = limpar_moeda(linhas[i + 1])
            if credito <= 0:
                i += 1; continue

            al = admin_raw.lower()
            if any(k in al for k in ('auto','automóvel','automovel','veículo','veiculo','carro','moto')):
                tipo = "Automóvel"
            elif any(k in al for k in ('imóvel','imovel','imov')):
                tipo = "Imóvel"
            elif any(k in al for k in ('caminhão','caminhao','pesado')):
                tipo = "Pesados"
            else:
                tipo = tipo_default

            if tipo_sel not in ("Todos","Geral") and tipo != tipo_sel:
                i += 2; continue

            admin = admin_raw.upper()
            for adm in _ADMINS:
                if adm.lower() in admin_raw.lower():
                    admin = adm.upper(); break

            entrada    = 0.0
            parcela    = 0.0
            n_parcelas = 0
            codigo_cota = None
            j = i + 2

            while j < min(i + 18, len(linhas)):
                lj   = linhas[j]
                lj_l = lj.lower()

                # Captura codigo da cota (3-4 digitos isolados)
                if re.match(r'^\d{3,4}$', lj) and codigo_cota is None:
                    codigo_cota = lj
                    j += 1; continue

                # Para na proxima admin
                if lj.lower() in _BANCOS_IC_LOWER:
                    break

                if lj_l == 'entrada:':
                    if j + 1 < len(linhas):
                        entrada = limpar_moeda(linhas[j + 1])
                        j += 2; continue

                if re.match(r'^parcelas?\s*:$', lj_l):
                    if j + 1 < len(linhas):
                        m2 = re.match(r'(\d+)\s*[xX]\s*R\$\s*([\d\.]+,\d{2})', linhas[j + 1])
                        if m2:
                            n_parcelas = int(m2.group(1))
                            parcela    = limpar_moeda(m2.group(2))
                    j += 2; continue

                m2 = re.search(r'entrada:\s*R\$\s*([\d\.]+,\d{2})', lj, re.IGNORECASE)
                if m2:
                    entrada = limpar_moeda(m2.group(1)); j += 1; continue

                m2 = re.search(r'parcelas?:\s*(\d+)\s*[xX]\s*R\$\s*([\d\.]+,\d{2})', lj, re.IGNORECASE)
                if m2:
                    n_parcelas = int(m2.group(1))
                    parcela    = limpar_moeda(m2.group(2))
                    j += 1; continue

                j += 1

            # Descarta se codigo esta na lista de reservados
            if codigo_cota and codigo_cota in codigos_reservados:
                i = j; continue

            if credito > 0 and entrada > 0:
                saldo = n_parcelas * parcela if n_parcelas > 0 and parcela > 0 \
                        else max(credito * 1.25 - entrada, credito * 0.20)
                lista.append({
                    'ID': id_c, 'Admin': admin, 'Tipo': tipo,
                    'Crédito': credito, 'Entrada': entrada,
                    'Parcela': parcela, 'NParcelas': n_parcelas,
                    'Saldo': saldo, 'CustoTotal': entrada + saldo,
                    'EntradaPct': entrada / credito,
                    'Disponivel': True,
                })
                id_c += 1

            i = j
            continue

        i += 1
    return lista


def _extrair_generico(texto: str, tipo_sel: str) -> list:
    blocos = re.split(
        r'(?i)(?=\b(?:imóvel|imovel|automóvel|automovel|veículo|veiculo|caminhão|caminhao|moto)\b)',
        texto
    )
    if len(blocos) < 2: blocos = re.split(r'\n\s*\n+', texto)
    if len(blocos) < 2: blocos = [texto]

    lista, id_c = [], 1
    for bloco in blocos:
        if len(bloco.strip()) < 20: continue
        try:
            bl = bloco.lower()
            admin = next((a.upper() for a in _ADMINS if a.lower() in bl), "OUTROS")
            if admin == "OUTROS" and "r$" not in bl: continue
            tipo = _detectar_tipo(bl)
            credito = 0.0
            m = _RE_CREDITO.search(bloco)
            if m: credito = limpar_moeda(m.group(1))
            if credito <= 0:
                vals = sorted([limpar_moeda(v) for v in _RE_MOEDA.findall(bloco) if limpar_moeda(v) > 5000], reverse=True)
                credito = vals[0] if vals else 0.0
            if credito <= 5000: continue
            if tipo == "Geral": tipo = "Imóvel" if credito > 250000 else "Automóvel"
            if tipo_sel not in ("Todos","Geral") and tipo != tipo_sel: continue
            entrada = 0.0
            m = _RE_ENTRADA.search(bloco)
            if m: entrada = limpar_moeda(m.group(1))
            if entrada <= 0:
                cands = sorted([limpar_moeda(v) for v in _RE_MOEDA.findall(bloco)
                                if credito * 0.01 < limpar_moeda(v) < credito * 0.95], reverse=True)
                entrada = cands[0] if cands else 0.0
            if entrada <= 0: continue
            saldo, parcela, n_parcelas = 0.0, 0.0, 0
            for pz_s, vl_s in _RE_PARCELA.findall(bloco):
                try:
                    pz, vl = int(pz_s), limpar_moeda(vl_s)
                    if pz > 0 and vl > 0:
                        saldo += pz * vl
                        if pz > 1 and vl > parcela:
                            parcela = vl; n_parcelas = pz
                except Exception: continue
            if saldo <= 0: saldo = max(credito * 1.25 - entrada, credito * 0.20)
            lista.append({
                'ID': id_c, 'Admin': admin, 'Tipo': tipo,
                'Crédito': credito, 'Entrada': entrada,
                'Parcela': parcela, 'NParcelas': n_parcelas,
                'Saldo': saldo, 'CustoTotal': entrada + saldo,
                'EntradaPct': entrada / credito,
            })
            id_c += 1
        except Exception: continue
    return lista


# ════════════════════════════════════════════════════════════
#  ROTEADOR PRINCIPAL
# ════════════════════════════════════════════════════════════
def extrair_dados_universal(texto: str, tipo_sel: str) -> list:
    if not texto or not texto.strip(): return []
    texto_limpo = "\n".join(
        ln.strip() for ln in texto.replace('\r\n','\n').replace('\r','\n').split('\n')
        if ln.strip()
    )
    fmt = _detectar_formato(texto_limpo)
    if fmt == "icontemplados_detalhe":
        return _extrair_icontemplados_detalhe(texto_limpo, tipo_sel)
    if fmt == "icontemplados_cards":
        return _extrair_icontemplados_cards(texto_limpo, tipo_sel)
    return _extrair_generico(texto_limpo, tipo_sel)


# ════════════════════════════════════════════════════════════
#  MOTOR
# ════════════════════════════════════════════════════════════
def _status(cr: float) -> str:
    if cr <= 0.20: return "💎 OURO"
    if cr <= 0.35: return "🔥 IMPERDÍVEL"
    if cr <= 0.45: return "✨ EXCELENTE"
    if cr <= 0.50: return "✅ OPORTUNIDADE"
    return "⚠️ PADRÃO"


def processar_combinacoes(cotas, min_cred, max_cred, max_ent, max_parc, max_custo, tipo_f, admin_f):
    MAX_ADM = 400
    TOL = 1.05
    filtradas = [c for c in cotas
                 if (tipo_f in ("Todos","Geral") or c['Tipo'] == tipo_f)
                 and (admin_f == "Todas" or c['Admin'] == admin_f)
                 and c['Admin'] != "OUTROS"
                 and c['Entrada'] <= max_ent * TOL
                 and c['Crédito'] <= max_cred
                 and c.get('Disponivel', True)]  # exclui cotas Reservadas
    if not filtradas: return pd.DataFrame()
    por_admin: dict = {}
    for c in filtradas: por_admin.setdefault(c['Admin'], []).append(c)
    res, prog, total = [], st.progress(0), len(por_admin)
    for idx, (admin, grupo) in enumerate(por_admin.items()):
        prog.progress(int((idx+1)/total*100))
        grupo.sort(key=lambda x: x['Entrada'])
        cnt = 0
        for r in range(1, 7):
            if cnt >= MAX_ADM: break
            for combo in itertools.combinations(grupo, r):
                if cnt >= MAX_ADM: break
                soma_e = sum(c['Entrada'] for c in combo)
                if soma_e > max_ent * TOL:
                    if r >= 2 and sum(c['Entrada'] for c in grupo[:r]) > max_ent * TOL: break
                    continue
                soma_c = sum(c['Crédito'] for c in combo)
                if soma_c < min_cred or soma_c > max_cred: continue
                soma_p = sum(c['Parcela'] for c in combo)
                if soma_p > max_parc * TOL: continue
                soma_s = sum(c['Saldo'] for c in combo)
                custo_t = soma_e + soma_s
                if soma_c <= 0: continue
                cr = (custo_t / soma_c) - 1
                if cr > max_custo: continue
                prazo = int(soma_s / soma_p) if soma_p > 0 else 0
                cet_mensal = (cr / prazo * 100) if prazo > 0 else 0.0
                res.append({
                    'STATUS': _status(cr), 'ADMINISTRADORA': admin,
                    'TIPO': combo[0]['Tipo'],
                    'IDS': " + ".join(str(c['ID']) for c in combo),
                    'CRÉDITO TOTAL': soma_c, 'ENTRADA TOTAL': soma_e,
                    'ENTRADA %': (soma_e/soma_c)*100, 'SALDO DEVEDOR': soma_s,
                    'CUSTO TOTAL': custo_t, 'PRAZO (meses)': prazo,
                    'PARCELA MENSAL': soma_p,
                    'CET TOTAL %': cr*100, 'CET MENSAL %': cet_mensal,
                    'DETALHES': " || ".join(
                        f"[ID {c['ID']}] {c['Admin']} · {fmt_brl(c['Crédito'])}" for c in combo),
                })
                cnt += 1
    prog.empty()
    if not res: return pd.DataFrame()
    return pd.DataFrame(res).sort_values('CET TOTAL %').reset_index(drop=True)


# ════════════════════════════════════════════════════════════
#  CACHE
# ════════════════════════════════════════════════════════════
@st.cache_data(max_entries=30, ttl=600, show_spinner=False)
def buscar_cached(_hash, texto, min_c, max_c, max_e, max_p, max_k, tipo, admin):
    cotas = extrair_dados_universal(texto, tipo)
    if not cotas: return pd.DataFrame()
    return processar_combinacoes(cotas, min_c, max_c, max_e, max_p, max_k, tipo, admin)

def gerar_hash(texto, *args):
    return hashlib.md5((texto + "|".join(str(a) for a in args)).encode()).hexdigest()


# ════════════════════════════════════════════════════════════
#  WHATSAPP
# ════════════════════════════════════════════════════════════
def gerar_msg_whatsapp(row: dict) -> str:
    tipo       = str(row.get('TIPO','Imóvel'))
    admin      = str(row.get('ADMINISTRADORA',''))
    credito    = float(row.get('CRÉDITO TOTAL', 0))
    entrada    = float(row.get('ENTRADA TOTAL', 0))
    parcela    = float(row.get('PARCELA MENSAL', 0))
    prazo      = int(row.get('PRAZO (meses)', 0))
    cet_total  = float(row.get('CET TOTAL %', 0))
    cet_mensal = float(row.get('CET MENSAL %', 0))
    ids        = str(row.get('IDS', ''))

    if "móvel" in tipo.lower() or "movel" in tipo.lower():
        emoji_tipo = "🏠 *IMÓVEL*"
    elif "pesado" in tipo.lower() or "caminhao" in tipo.lower():
        emoji_tipo = "🚛 *CAMINHÃO*"
    else:
        emoji_tipo = "🚗 *AUTO*"

    parc_str = f"{prazo}x {fmt_brl(parcela)}" if prazo > 0 and parcela > 0 else "A consultar"

    # Quantidade de cotas na junção (ex: "1 + 56" → 2 cotas)
    n_cotas = len(ids.split('+')) if ids else 1

    # Taxa de transferência: Itaú = R$ 650 por cota | demais = 1% do crédito total
    admin_upper = admin.upper()
    if 'ITAÚ' in admin_upper or 'ITAU' in admin_upper:
        tx_transf = 650.0 * n_cotas
    else:
        tx_transf = credito * 0.01

    return (
        f"🔑 *CARTA CONTEMPLADA — {admin}*\n"
        f"\n"
        f"{emoji_tipo}\n"
        f"*Crédito:* {fmt_brl(credito)}\n"
        f"*Entrada:* {fmt_brl(entrada)}\n"
        f"*Parcela:* {parc_str}\n"
        f"*Tx de Transferência:* {fmt_brl(tx_transf)}\n"
        f"\n"
        f"*CET Mensal:* {fmt_pct(cet_mensal)} a.m.\n"
        f"*CET Total:* {fmt_pct(cet_total)}\n"
        f"\n"
        f"━━━━━━━━━━━━━━━━━\n"
        f"🏆 *JBS Contempladas*"
    )


# ════════════════════════════════════════════════════════════
#  PDF
# ════════════════════════════════════════════════════════════
def _san(t) -> str:
    return str(t).encode('latin-1','replace').decode('latin-1')

class RelatorioPDF(FPDF):
    def header(self):
        self.set_fill_color(132,117,78); self.rect(0,0,297,20,'F')
        if os.path.exists("logo_pdf.png"): self.image('logo_pdf.png',4,2,30)
        self.set_font('Arial','B',14); self.set_text_color(255,255,255)
        self.set_xy(38,5); self.cell(0,10,'JBS SNIPER  |  RELATORIO DE OPORTUNIDADES',0,1,'L')
        self.set_font('Arial','',7); self.set_xy(38,13)
        self.cell(0,5,f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}',0,1,'L'); self.ln(4)
    def footer(self):
        self.set_y(-12); self.set_font('Arial','I',7); self.set_text_color(150,150,150)
        self.cell(0,8,f'JBS SNIPER  |  Pagina {self.page_no()}',0,0,'C')

def gerar_pdf(df: pd.DataFrame) -> bytes:
    pdf = RelatorioPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()
    pdf.set_fill_color(236,236,228); pdf.set_text_color(30,30,30); pdf.set_font('Arial','B',7)
    cols = ["STS","ADMIN","TIPO","CREDITO","ENTRADA","ENT%","SALDO","CUSTO TOT","PRZ","PARCELA","CET TOT","CET MEN","DETALHES"]
    wids = [16,    18,     12,    26,        26,       8,     26,     26,         7,    22,        11,       11,       47]
    for h,w in zip(cols,wids): pdf.cell(w,7,h,1,0,'C',True)
    pdf.ln(); pdf.set_font('Arial','',6.5)
    for i,(_,row) in enumerate(df.iterrows()):
        fill=(i%2==0)
        pdf.set_fill_color(245,245,240) if fill else pdf.set_fill_color(255,255,255)
        pdf.cell(wids[0],  6, _san(str(row['STATUS']))[:6],          1,0,'C',fill)
        pdf.cell(wids[1],  6, _san(str(row['ADMINISTRADORA']))[:12], 1,0,'C',fill)
        pdf.cell(wids[2],  6, _san(str(row['TIPO']))[:8],            1,0,'C',fill)
        pdf.cell(wids[3],  6, _san(fmt_brl(row['CRÉDITO TOTAL'])),   1,0,'R',fill)
        pdf.cell(wids[4],  6, _san(fmt_brl(row['ENTRADA TOTAL'])),   1,0,'R',fill)
        pdf.cell(wids[5],  6, _san(fmt_pct_curto(row['ENTRADA %'])), 1,0,'C',fill)
        pdf.cell(wids[6],  6, _san(fmt_brl(row['SALDO DEVEDOR'])),   1,0,'R',fill)
        pdf.cell(wids[7],  6, _san(fmt_brl(row['CUSTO TOTAL'])),     1,0,'R',fill)
        pdf.cell(wids[8],  6, str(int(row['PRAZO (meses)'])),        1,0,'C',fill)
        pdf.cell(wids[9],  6, _san(fmt_brl(row['PARCELA MENSAL'])),  1,0,'R',fill)
        pdf.cell(wids[10], 6, _san(fmt_pct(row['CET TOTAL %'])),     1,0,'C',fill)
        pdf.cell(wids[11], 6, _san(fmt_pct(row['CET MENSAL %'])),    1,0,'C',fill)
        pdf.cell(wids[12], 6, _san(str(row['DETALHES']))[:42],       1,1,'L',fill)
    out = pdf.output(dest='S')
    return bytes(out) if isinstance(out,(bytes,bytearray)) else out.encode('latin-1')


# ════════════════════════════════════════════════════════════
#  EXCEL
# ════════════════════════════════════════════════════════════
def gerar_excel(df: pd.DataFrame, cotas: list) -> bytes:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    GOLD_HEX, ZEBRA = "84754e", "f5f5f0"
    FMT_BRL = '_-"R$"* #.##0,00_-;-"R$"* #.##0,00_-;_-"R$"* "-"??_-;_-@_-'
    FMT_PCT = '0,00"%"'
    FMT_NUM = '#.##0'
    borda = Border(left=Side(style='thin',color='CCCCCC'), right=Side(style='thin',color='CCCCCC'),
                   top=Side(style='thin',color='CCCCCC'),  bottom=Side(style='thin',color='CCCCCC'))
    col_names = list(df.columns)
    col_widths = {
        'STATUS':18,'ADMINISTRADORA':16,'TIPO':12,'IDS':24,
        'CRÉDITO TOTAL':22,'ENTRADA TOTAL':22,'ENTRADA %':12,
        'SALDO DEVEDOR':22,'CUSTO TOTAL':22,'PRAZO (meses)':14,
        'PARCELA MENSAL':22,'CET TOTAL %':14,'CET MENSAL %':14,'DETALHES':55,
    }
    fmt_map = {}
    for i,col in enumerate(col_names, start=1):
        if col in ('CRÉDITO TOTAL','ENTRADA TOTAL','SALDO DEVEDOR','CUSTO TOTAL','PARCELA MENSAL'): fmt_map[i]=FMT_BRL
        elif col in ('ENTRADA %','CET TOTAL %','CET MENSAL %'): fmt_map[i]=FMT_PCT
        elif col=='PRAZO (meses)': fmt_map[i]=FMT_NUM
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Oportunidades')
        ws = writer.sheets['Oportunidades']
        for cell in ws[1]:
            cell.fill=PatternFill("solid",fgColor=GOLD_HEX); cell.font=Font(bold=True,color="FFFFFF",size=10)
            cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=borda
            ws.column_dimensions[get_column_letter(cell.column)].width=col_widths.get(col_names[cell.column-1],14)
        ws.row_dimensions[1].height=22
        for ri,row in enumerate(ws.iter_rows(min_row=2,max_row=ws.max_row),start=2):
            z=(ri%2==0)
            for cell in row:
                cell.border=borda; cell.font=Font(size=9)
                if z: cell.fill=PatternFill("solid",fgColor=ZEBRA)
                ci=cell.column; cn=col_names[ci-1]
                if ci in fmt_map:
                    cell.number_format=fmt_map[ci]
                    try: cell.value=float(cell.value) if cell.value not in (None,'') else cell.value
                    except: pass
                    cell.alignment=Alignment(horizontal='right',vertical='center')
                elif cn in ('DETALHES','IDS'): cell.alignment=Alignment(horizontal='left',vertical='center')
                else: cell.alignment=Alignment(horizontal='center',vertical='center')
        ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
        if cotas:
            df_c=pd.DataFrame(cotas).drop(columns=['NParcelas'],errors='ignore')
            df_c.to_excel(writer,index=False,sheet_name='Cotas Lidas')
            ws2=writer.sheets['Cotas Lidas']
            cw={'ID':6,'Admin':16,'Tipo':12,'Crédito':22,'Entrada':22,'Parcela':22,'Saldo':22,'CustoTotal':22,'EntradaPct':14}
            cf={'Crédito':FMT_BRL,'Entrada':FMT_BRL,'Parcela':FMT_BRL,'Saldo':FMT_BRL,'CustoTotal':FMT_BRL,'EntradaPct':FMT_PCT}
            cc=list(df_c.columns)
            for cell in ws2[1]:
                cell.fill=PatternFill("solid",fgColor=GOLD_HEX); cell.font=Font(bold=True,color="FFFFFF",size=10)
                cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=borda
                ws2.column_dimensions[get_column_letter(cell.column)].width=cw.get(cc[cell.column-1],14)
            for ri,row in enumerate(ws2.iter_rows(min_row=2,max_row=ws2.max_row),start=2):
                for cell in row:
                    cell.border=borda; cell.font=Font(size=9)
                    cell.alignment=Alignment(horizontal='center',vertical='center')
                    if ri%2==0: cell.fill=PatternFill("solid",fgColor=ZEBRA)
                    cn=cc[cell.column-1]
                    if cn in cf:
                        cell.number_format=cf[cn]
                        try: cell.value=float(cell.value) if cell.value not in (None,'') else cell.value
                        except: pass
                        cell.alignment=Alignment(horizontal='right',vertical='center')
            ws2.freeze_panes='A2'
    return buf.getvalue()


# ════════════════════════════════════════════════════════════
#  PDF PREMIUM — COTA CONTEMPLADA INDIVIDUAL
# ════════════════════════════════════════════════════════════
def gerar_pdf_contemplada(
    admin: str, tipo: str, nome_cliente: str,
    credito: float, entrada: float, n_parcelas: int, parcela: float,
    tx_transf: float
) -> bytes:
    """
    Gera PDF premium de apresentação de cota contemplada individual.
    Layout: capa com dados financeiros + página CTA de fechamento.
    """
    saldo_devedor = n_parcelas * parcela
    custo_total   = entrada + saldo_devedor + tx_transf
    cet_total_pct = ((custo_total / credito) - 1) * 100 if credito > 0 else 0
    cet_mensal_pct = cet_total_pct / n_parcelas if n_parcelas > 0 else 0

    if "vel" in tipo.lower() or "auto" in tipo.lower():
        tipo_label = "AUTOMOVEL"
        tipo_emoji_txt = "Veiculo / Automovel"
    elif "movel" in tipo.lower() or "vel" in tipo.lower():
        tipo_label = "IMOVEL"
        tipo_emoji_txt = "Imovel / Bem de Raiz"
    elif "pesado" in tipo.lower() or "caminhao" in tipo.lower():
        tipo_label = "PESADOS"
        tipo_emoji_txt = "Caminhao / Pesados"
    else:
        tipo_label = tipo.upper()
        tipo_emoji_txt = tipo

    GOLD_R, GOLD_G, GOLD_B   = 132, 117, 78
    DARK_R, DARK_G, DARK_B   = 14,  17,  23
    WHITE_R, WHITE_G, WHITE_B = 255, 255, 255
    LGRAY_R, LGRAY_G, LGRAY_B = 245, 244, 238
    GREEN_R, GREEN_G, GREEN_B = 39, 174, 96

    def _s(t): return str(t).encode('latin-1','replace').decode('latin-1')

    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()

    # ── Fundo escuro topo ──────────────────────────────────
    pdf.set_fill_color(DARK_R, DARK_G, DARK_B)
    pdf.rect(0, 0, 210, 68, 'F')

    # ── Logo área ──────────────────────────────────────────
    if os.path.exists("logo_pdf.png"):
        pdf.image("logo_pdf.png", 12, 8, 36)

    # ── Título principal ───────────────────────────────────
    pdf.set_text_color(GOLD_R, GOLD_G, GOLD_B)
    pdf.set_font('Arial', 'B', 26)
    pdf.set_xy(55, 12)
    pdf.cell(0, 10, _s("CARTA CONTEMPLADA"), 0, 1, 'L')

    pdf.set_font('Arial', '', 11)
    pdf.set_text_color(WHITE_R, WHITE_G, WHITE_B)
    pdf.set_xy(55, 24)
    pdf.cell(0, 6, _s(f"Administradora: {admin.upper()}"), 0, 1, 'L')

    pdf.set_font('Arial', 'B', 11)
    pdf.set_text_color(GOLD_R, GOLD_G, GOLD_B)
    pdf.set_xy(55, 31)
    pdf.cell(0, 6, _s(tipo_emoji_txt.upper()), 0, 1, 'L')

    # ── Data ──────────────────────────────────────────────
    pdf.set_font('Arial', '', 8)
    pdf.set_text_color(180, 170, 130)
    pdf.set_xy(55, 38)
    pdf.cell(0, 5, _s(f"Emitido em {datetime.now().strftime('%d/%m/%Y  |  Validade: consultar disponibilidade')}"), 0, 1, 'L')

    # ── Nome cliente ──────────────────────────────────────
    if nome_cliente.strip():
        pdf.set_fill_color(GOLD_R, GOLD_G, GOLD_B)
        pdf.rect(55, 46, 143, 14, 'F')
        pdf.set_font('Arial', 'B', 12)
        pdf.set_text_color(WHITE_R, WHITE_G, WHITE_B)
        pdf.set_xy(58, 49)
        pdf.cell(0, 7, _s(f"Preparado para: {nome_cliente.upper()}"), 0, 1, 'L')

    # ── Linha dourada separadora ───────────────────────────
    pdf.set_fill_color(GOLD_R, GOLD_G, GOLD_B)
    pdf.rect(0, 68, 210, 2, 'F')

    # ══ BLOCO FINANCEIRO — 3 colunas ═════════════════════
    y_bloco = 76
    w_col   = 62
    gap     = 4
    x_cols  = [12, 12 + w_col + gap, 12 + 2*(w_col + gap)]

    def card_financeiro(x, y, titulo, valor_str, cor_valor=(14,17,23), destaque=False):
        if destaque:
            pdf.set_fill_color(GOLD_R, GOLD_G, GOLD_B)
            pdf.rect(x, y, w_col, 26, 'F')
            pdf.set_text_color(WHITE_R, WHITE_G, WHITE_B)
        else:
            pdf.set_fill_color(LGRAY_R, LGRAY_G, LGRAY_B)
            pdf.rect(x, y, w_col, 26, 'F')
            pdf.set_text_color(90, 80, 50)
        pdf.set_font('Arial', '', 8)
        pdf.set_xy(x+2, y+3)
        pdf.cell(w_col-4, 5, _s(titulo.upper()), 0, 1, 'C')
        if destaque:
            pdf.set_text_color(WHITE_R, WHITE_G, WHITE_B)
        else:
            pdf.set_text_color(*cor_valor)
        pdf.set_font('Arial', 'B', 14)
        pdf.set_xy(x+2, y+10)
        pdf.cell(w_col-4, 10, _s(valor_str), 0, 1, 'C')

    card_financeiro(x_cols[0], y_bloco, "CREDITO DA CARTA",
                    fmt_brl(credito), destaque=True)
    card_financeiro(x_cols[1], y_bloco, "ENTRADA (investimento)",
                    fmt_brl(entrada), cor_valor=(14,17,23))
    card_financeiro(x_cols[2], y_bloco, "TX DE TRANSFERENCIA",
                    fmt_brl(tx_transf), cor_valor=(14,17,23))

    y_bloco2 = y_bloco + 32
    card_financeiro(x_cols[0], y_bloco2, "SALDO DEVEDOR",
                    fmt_brl(saldo_devedor), cor_valor=(14,17,23))
    card_financeiro(x_cols[1], y_bloco2, "PARCELAS",
                    f"{n_parcelas}x {fmt_brl(parcela)}", cor_valor=(14,17,23))
    card_financeiro(x_cols[2], y_bloco2, "CUSTO TOTAL DA OPERACAO",
                    fmt_brl(custo_total), destaque=True)

    # ══ BLOCO CET ══════════════════════════════════════
    y_cet = y_bloco2 + 33
    pdf.set_fill_color(DARK_R, DARK_G, DARK_B)
    pdf.rect(12, y_cet, 186, 28, 'F')

    pdf.set_text_color(GOLD_R, GOLD_G, GOLD_B)
    pdf.set_font('Arial', 'B', 10)
    pdf.set_xy(14, y_cet + 4)
    pdf.cell(60, 6, _s("CET MENSAL (a.m.)"), 0, 0, 'L')
    pdf.set_xy(80, y_cet + 4)
    pdf.cell(60, 6, _s("CET TOTAL"), 0, 0, 'L')
    pdf.set_xy(146, y_cet + 4)
    pdf.cell(50, 6, _s("PRAZO"), 0, 0, 'L')

    pdf.set_text_color(WHITE_R, WHITE_G, WHITE_B)
    pdf.set_font('Arial', 'B', 18)
    pdf.set_xy(14, y_cet + 11)
    pdf.cell(60, 10, _s(fmt_pct(cet_mensal_pct)), 0, 0, 'L')
    pdf.set_xy(80, y_cet + 11)
    pdf.cell(60, 10, _s(fmt_pct(cet_total_pct)), 0, 0, 'L')
    pdf.set_font('Arial', 'B', 14)
    pdf.set_xy(146, y_cet + 11)
    pdf.cell(50, 10, _s(f"{n_parcelas} meses"), 0, 0, 'L')

    # ══ LINHA DOURADA ══════════════════════════════════
    y_sep = y_cet + 33
    pdf.set_fill_color(GOLD_R, GOLD_G, GOLD_B)
    pdf.rect(12, y_sep, 186, 1, 'F')

    # ══ COMO FUNCIONA — 3 passos ══════════════════════
    y_steps = y_sep + 6
    pdf.set_text_color(DARK_R, DARK_G, DARK_B)
    pdf.set_font('Arial', 'B', 12)
    pdf.set_xy(12, y_steps)
    pdf.cell(186, 7, _s("COMO FUNCIONA"), 0, 1, 'L')

    steps = [
        ("1. VOCE INVESTE", f"Entrada de {fmt_brl(entrada)} para transferir a cota ja contemplada para seu nome."),
        ("2. CREDITO LIBERADO", f"Credito de {fmt_brl(credito)} disponivel imediatamente para uso — imovel ou veiculo."),
        ("3. PARCELAS NO PRAZO", f"{n_parcelas} parcelas de {fmt_brl(parcela)} ate o encerramento do grupo.")
    ]
    y_steps += 9
    for titulo_step, desc_step in steps:
        pdf.set_fill_color(LGRAY_R, LGRAY_G, LGRAY_B)
        pdf.rect(12, y_steps, 186, 16, 'F')
        pdf.set_fill_color(GOLD_R, GOLD_G, GOLD_B)
        pdf.rect(12, y_steps, 3, 16, 'F')
        pdf.set_text_color(DARK_R, DARK_G, DARK_B)
        pdf.set_font('Arial', 'B', 9)
        pdf.set_xy(18, y_steps + 2)
        pdf.cell(80, 5, _s(titulo_step), 0, 1, 'L')
        pdf.set_font('Arial', '', 8)
        pdf.set_text_color(60, 55, 40)
        pdf.set_xy(18, y_steps + 8)
        pdf.cell(178, 5, _s(desc_step), 0, 1, 'L')
        y_steps += 19

    # ══ CTA BOX — fechamento ══════════════════════════
    y_cta = y_steps + 4
    pdf.set_fill_color(GREEN_R, GREEN_G, GREEN_B)
    pdf.rect(12, y_cta, 186, 36, 'F')

    pdf.set_text_color(WHITE_R, WHITE_G, WHITE_B)
    pdf.set_font('Arial', 'B', 14)
    pdf.set_xy(14, y_cta + 4)
    pdf.cell(182, 8, _s("ESTA COTA ESTA DISPONIVEL — VAGAS LIMITADAS"), 0, 1, 'C')

    pdf.set_font('Arial', '', 9)
    pdf.set_xy(14, y_cta + 13)
    pdf.multi_cell(182, 5, _s(
        "Cartas contempladas sao transferidas rapidamente. Voce ja tem o credito na mao — "
        "sem sorteio, sem fila. Basta formalizar e o credito e seu.\n"
        "Entre em contato agora e reserve esta oportunidade antes que outro cliente feche."
    ), 0, 'C')

    pdf.set_font('Arial', 'B', 11)
    pdf.set_xy(14, y_cta + 28)
    pdf.cell(182, 6, _s("Fale com nosso consultor: JBS Contempladas"), 0, 1, 'C')

    # ══ RODAPE ══════════════════════════════════════
    pdf.set_fill_color(DARK_R, DARK_G, DARK_B)
    pdf.rect(0, 280, 210, 17, 'F')
    pdf.set_text_color(GOLD_R, GOLD_G, GOLD_B)
    pdf.set_font('Arial', 'B', 8)
    pdf.set_xy(12, 283)
    pdf.cell(186, 5, _s("JBS CONTEMPLADAS  |  Inteligencia Comercial  |  Documento gerado automaticamente"), 0, 1, 'C')
    pdf.set_text_color(130, 120, 80)
    pdf.set_font('Arial', '', 7)
    pdf.set_xy(12, 289)
    pdf.cell(186, 4, _s("Valores sujeitos a confirmacao. Consulte disponibilidade com nossa equipe antes de qualquer decisao."), 0, 1, 'C')

    out = pdf.output(dest='S')
    return out if isinstance(out, bytes) else out.encode('latin-1')


# ════════════════════════════════════════════════════════════
#  SESSION STATE
# ════════════════════════════════════════════════════════════
for k,v in [('df_resultado',None),('admins_disponiveis',["Todas"]),
            ('cotas_lidas',[]),('msg_whatsapp',"")]:
    if k not in st.session_state: st.session_state[k]=v


# ════════════════════════════════════════════════════════════
#  ABAS PRINCIPAIS
# ════════════════════════════════════════════════════════════
tab_sniper, tab_contemplada = st.tabs([
    "🎯  SNIPER — Busca de Cotas",
    "🔑  CONTEMPLADA — PDF Individual",
])


# ════════════════════════════════════════════════════════════
#  ABA 1 — SNIPER (busca combinatória)
# ════════════════════════════════════════════════════════════
with tab_sniper:
    with st.expander("📋  DADOS DO SITE  —  Cole o texto das cotas aqui", expanded=True):
        texto_site = st.text_area("Texto copiado do portal / WhatsApp", height=130,
                                  placeholder="Cole aqui o texto com as cotas de consórcio...", key="input_texto")
        if texto_site:
            prev = extrair_dados_universal(texto_site, "Todos")
            st.session_state.cotas_lidas = prev
            st.session_state.admins_disponiveis = ["Todas"] + sorted(
                set(c['Admin'] for c in prev if c['Admin'] != "OUTROS"))
            if prev:
                tp = {}
                for c in prev: tp[c['Tipo']] = tp.get(c['Tipo'],0)+1
                disponiveis = sum(1 for c in prev if c.get('Disponivel', True))
                reservadas  = len(prev) - disponiveis
                resumo = ' · '.join(f'{v} {k}' for k,v in sorted(tp.items()))
                aviso_res = f" _(⚠️ {reservadas} reservadas excluídas)_" if reservadas > 0 else ""
                st.success(f"✅  **{disponiveis} cotas disponíveis** de {len(prev)} lidas — {resumo}{aviso_res}")
            else:
                st.warning("⚠️  Nenhuma cota identificada.")
        else:
            st.session_state.cotas_lidas = []
            st.session_state.admins_disponiveis = ["Todas"]


    # ════════════════════════════════════════════════════════════
    #  FILTROS
    # ════════════════════════════════════════════════════════════
    st.markdown(f"<h3 style='margin-top:8px'>⚙️ FILTROS DE BUSCA</h3>", unsafe_allow_html=True)
    r1c1,r1c2,r1c3,r1c4 = st.columns(4)
    tipo_bem     = r1c1.selectbox("Tipo de bem",    ["Todos","Imóvel","Automóvel","Pesados"])
    admin_filtro = r1c2.selectbox("Administradora", st.session_state.admins_disponiveis)
    min_c = r1c3.number_input("Crédito mínimo (R$)", 0.0, step=1000.0, value=60_000.0,  format="%.2f")
    max_c = r1c4.number_input("Crédito máximo (R$)", 0.0, step=1000.0, value=710_000.0, format="%.2f")
    r2c1,r2c2,r2c3 = st.columns(3)
    max_e = r2c1.number_input("Entrada máxima (R$)", 0.0, step=1000.0, value=200_000.0, format="%.2f")
    max_p = r2c2.number_input("Parcela máxima (R$)", 0.0, step=100.0,  value=4_500.0,   format="%.2f")
    max_k = r2c3.slider("Custo efetivo máximo (%)", 0.0, 100.0, 55.0, 0.5, format="%.1f%%")
    max_k_dec = max_k / 100.0
    st.markdown("")

    cb,ci = st.columns([2,3])
    with cb: buscar = st.button("🔍  LOCALIZAR OPORTUNIDADES", type="primary")
    with ci:
        if st.session_state.cotas_lidas:
            n = len(st.session_state.cotas_lidas)
            st.markdown(f"<p style='color:{GOLD};font-size:.85rem;margin-top:10px'>🎯 {n} cota{'s' if n!=1 else ''} disponível{'is' if n!=1 else ''} para análise</p>", unsafe_allow_html=True)

    if buscar:
        if not texto_site:
            st.error("❌  Cole os dados das cotas antes de buscar.")
        elif not st.session_state.cotas_lidas:
            st.error("❌  Nenhuma cota identificada no texto.")
        else:
            with st.spinner("🔍  Analisando combinações..."):
                _h = gerar_hash(texto_site, min_c, max_c, max_e, max_p, max_k_dec, tipo_bem, admin_filtro)
                df = buscar_cached(_h, texto_site, min_c, max_c, max_e, max_p, max_k_dec, tipo_bem, admin_filtro)
            st.session_state.df_resultado = df
            st.session_state.msg_whatsapp = ""


    # ════════════════════════════════════════════════════════════
    #  RESULTADOS
    # ════════════════════════════════════════════════════════════
    if st.session_state.df_resultado is not None:
        df_show = st.session_state.df_resultado
        if df_show.empty:
            st.warning("🔎  Nenhuma combinação encontrada. Tente ampliar os filtros.")
        else:
            st.markdown(f"<h3 style='margin-top:24px'>📊 RESULTADO DA ANÁLISE</h3>", unsafe_allow_html=True)
            m1,m2,m3,m4,m5 = st.columns(5)
            m1.metric("Oportunidades",     str(len(df_show)))
            m2.metric("Menor CET",         fmt_pct(df_show['CET TOTAL %'].min()))
            m3.metric("Menor entrada",     fmt_brl(df_show['ENTRADA TOTAL'].min()))
            m4.metric("Maior crédito",     fmt_brl(df_show['CRÉDITO TOTAL'].max()))
            m5.metric("💎 Ouro/Imperdível", str(len(df_show[df_show['STATUS'].str.contains("OURO|IMPERDÍVEL")])))
            st.markdown("")

            df_exib = df_show.copy()
            for col in ('CRÉDITO TOTAL','ENTRADA TOTAL','SALDO DEVEDOR','CUSTO TOTAL','PARCELA MENSAL'):
                df_exib[col] = df_exib[col].apply(fmt_brl)
            df_exib['ENTRADA %']    = df_exib['ENTRADA %'].apply(fmt_pct_curto)
            df_exib['CET TOTAL %']  = df_exib['CET TOTAL %'].apply(fmt_pct)
            df_exib['CET MENSAL %'] = df_exib['CET MENSAL %'].apply(fmt_pct)

            cols_tab = [c for c in df_exib.columns if c != 'DETALHES']
            st.dataframe(df_exib[cols_tab], hide_index=True, use_container_width=True, height=400)

            with st.expander("🔎  Ver detalhes das combinações"):
                st.dataframe(df_exib[['STATUS','ADMINISTRADORA','IDS','CET TOTAL %','DETALHES']],
                             hide_index=True, use_container_width=True)

            if st.session_state.cotas_lidas:
                with st.expander("📋  Ver cotas lidas individualmente"):
                    df_c = pd.DataFrame(st.session_state.cotas_lidas).drop(columns=['NParcelas'],errors='ignore').copy()
                    for col in ('Crédito','Entrada','Parcela','Saldo','CustoTotal'):
                        df_c[col] = df_c[col].apply(fmt_brl)
                    df_c['EntradaPct'] = df_c['EntradaPct'].apply(fmt_pct)
                    st.dataframe(df_c, hide_index=True, use_container_width=True)

            # WHATSAPP
            st.markdown(f"<h3 style='margin-top:28px'>📲 GERAR MENSAGEM WHATSAPP</h3>", unsafe_allow_html=True)
            st.markdown(f"<p style='color:{BEIGE};opacity:.55;font-size:.83rem;margin-bottom:10px'>Clique em uma linha para selecionar, depois clique em GERAR MENSAGEM.</p>", unsafe_allow_html=True)
            cols_sel = ['STATUS','ADMINISTRADORA','TIPO','IDS','CRÉDITO TOTAL','ENTRADA TOTAL','PARCELA MENSAL','CET TOTAL %','CET MENSAL %']
            df_sel = df_exib[[c for c in cols_sel if c in df_exib.columns]].copy()
            df_sel.insert(0,'Nº', range(1, len(df_sel)+1))
            sel = st.dataframe(df_sel, hide_index=True, use_container_width=True, height=260,
                               on_select="rerun", selection_mode="single-row")
            linha_sel = None
            if sel and hasattr(sel,'selection') and sel.selection and sel.selection.get('rows'):
                linha_sel = sel.selection['rows'][0]
            col_btn_w, col_info_w = st.columns([2,3])
            with col_btn_w:
                gerar_msg = st.button("📲  GERAR MENSAGEM", type="primary")
            with col_info_w:
                if linha_sel is not None:
                    rp = df_show.iloc[linha_sel]
                    st.markdown(f"<p style='color:{GOLD};font-size:.85rem;margin-top:10px'>✅ Linha {linha_sel+1} — {rp['ADMINISTRADORA']} · {fmt_brl(rp['CRÉDITO TOTAL'])}</p>", unsafe_allow_html=True)
                else:
                    st.markdown(f"<p style='color:{BEIGE};opacity:.35;font-size:.85rem;margin-top:10px'>Nenhuma linha selecionada</p>", unsafe_allow_html=True)
            if gerar_msg:
                if linha_sel is None:
                    st.warning("⚠️  Selecione uma linha antes de gerar.")
                else:
                    st.session_state.msg_whatsapp = gerar_msg_whatsapp(df_show.iloc[linha_sel].to_dict())
            if st.session_state.msg_whatsapp:
                st.markdown(f"<p style='color:{GOLD};font-weight:600;margin-top:16px'>📋 Copie abaixo e cole no WhatsApp:</p>", unsafe_allow_html=True)
                st.code(st.session_state.msg_whatsapp, language=None)

            # DOWNLOADS
            st.markdown(f"<h3 style='margin-top:28px'>⬇️ EXPORTAR</h3>", unsafe_allow_html=True)
            dl2, dl3 = st.columns(2)
            ts = datetime.now().strftime('%Y%m%d_%H%M')
            try:
                xl_b = gerar_excel(df_show, st.session_state.cotas_lidas)
                dl2.download_button("📊  Baixar Excel", data=xl_b,
                                    file_name=f"sniper_{ts}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True)
            except Exception as ex:
                dl2.error(f"Excel: {ex}")
            try:
                pdf_b = gerar_pdf(df_show.head(200))
                dl3.download_button("📑  Baixar PDF", data=pdf_b,
                                    file_name=f"sniper_{ts}.pdf", mime="application/pdf",
                                    use_container_width=True)
            except Exception as ex:
                dl3.error(f"PDF: {ex}")


# ════════════════════════════════════════════════════════════
#  ABA 2 — CONTEMPLADA — PDF individual
# ════════════════════════════════════════════════════════════
with tab_contemplada:
    st.markdown(f"<h3 style='margin-top:8px'>🔑 GERAR PDF DE COTA CONTEMPLADA</h3>", unsafe_allow_html=True)
    st.markdown(
        f"<p style='color:{BEIGE};opacity:.65;font-size:.88rem;margin-bottom:16px'>"
        "Preencha os dados da cota e gere um PDF premium para enviar ao cliente.</p>",
        unsafe_allow_html=True)

    cc1, cc2 = st.columns([3, 2])
    with cc1:
        ct_admin      = st.text_input("Administradora", value="REMAZA", key="ct_admin")
        ct_tipo       = st.selectbox("Tipo de bem", ["Imóvel","Automóvel","Pesados"], key="ct_tipo")
        ct_nome       = st.text_input("Nome do cliente (opcional)", value="", key="ct_nome",
                                      placeholder="Ex: João Silva")
        ct_credito    = st.number_input("Crédito da carta (R$)", 0.0, step=1000.0,
                                        value=225_000.0, format="%.2f", key="ct_credito")
        ct_entrada    = st.number_input("Entrada / Investimento (R$)", 0.0, step=500.0,
                                        value=94_500.0, format="%.2f", key="ct_entrada")
        ct_n_parcelas = st.number_input("Nº de parcelas", 1, 360, value=156, step=1, key="ct_nparcelas")
        ct_parcela    = st.number_input("Valor da parcela (R$)", 0.0, step=50.0,
                                        value=1_419.33, format="%.2f", key="ct_parcela")

    with cc2:
        st.markdown(f"<p style='color:{GOLD};font-size:.78rem;font-weight:600;letter-spacing:1px;text-transform:uppercase;margin-top:8px'>Taxa de transferência</p>", unsafe_allow_html=True)
        tx_opcao = st.radio("Cálculo da taxa",
                            ["Auto (1% do crédito)","Itaú (R$ 650)","Manual"],
                            key="ct_tx_opcao", label_visibility="collapsed")
        if tx_opcao == "Manual":
            ct_tx = st.number_input("Valor da taxa (R$)", 0.0, step=100.0, value=0.0,
                                    format="%.2f", key="ct_tx_manual")
        elif tx_opcao == "Itaú (R$ 650)":
            ct_tx = 650.0
            st.info("Taxa Itaú: R$ 650,00")
        else:
            ct_tx = ct_credito * 0.01
            st.info(f"Taxa automática (1%): {fmt_brl(ct_tx)}")

        if ct_credito > 0 and ct_entrada > 0:
            saldo_prev = ct_n_parcelas * ct_parcela
            custo_prev = ct_entrada + saldo_prev + ct_tx
            cet_tot    = ((custo_prev / ct_credito) - 1) * 100
            cet_men    = cet_tot / ct_n_parcelas if ct_n_parcelas > 0 else 0
            preview_html = (
                f"<div style='background:{CARD};border:1px solid {BORDER};border-radius:10px;padding:16px;margin-top:12px'>"
                f"<p style='color:{GOLD};font-size:.72rem;letter-spacing:1.5px;margin:0 0 8px;font-weight:700'>PRÉ-VISUALIZAÇÃO</p>"
                f"<p style='margin:2px 0;font-size:.9rem'><b>Saldo devedor:</b> {fmt_brl(saldo_prev)}</p>"
                f"<p style='margin:2px 0;font-size:.9rem'><b>Custo total:</b> {fmt_brl(custo_prev)}</p>"
                f"<p style='margin:2px 0;font-size:.9rem'><b>CET Total:</b> {fmt_pct(cet_tot)}</p>"
                f"<p style='margin:4px 0 0;font-size:.9rem'><b>CET Mensal:</b> {fmt_pct(cet_men)} a.m.</p>"
                "</div>"
            )
            st.markdown(preview_html, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    col_gerar, _ = st.columns([2, 3])
    with col_gerar:
        gerar_ct = st.button("📄  GERAR PDF PREMIUM", type="primary", key="btn_gerar_ct")

    if gerar_ct:
        if ct_credito <= 0 or ct_entrada <= 0 or ct_parcela <= 0:
            st.error("❌  Preencha pelo menos Crédito, Entrada e Parcela.")
        else:
            with st.spinner("Gerando PDF premium..."):
                try:
                    pdf_ct = gerar_pdf_contemplada(
                        admin=ct_admin, tipo=ct_tipo, nome_cliente=ct_nome,
                        credito=ct_credito, entrada=ct_entrada,
                        n_parcelas=int(ct_n_parcelas), parcela=ct_parcela,
                        tx_transf=ct_tx
                    )
                    ts_ct = datetime.now().strftime('%Y%m%d_%H%M')
                    nome_arquivo = f"contemplada_{ct_admin.lower().replace(' ','_')}_{ts_ct}.pdf"
                    st.success("✅  PDF gerado com sucesso!")
                    st.download_button(
                        label="⬇️  BAIXAR PDF — CONTEMPLADA",
                        data=pdf_ct,
                        file_name=nome_arquivo,
                        mime="application/pdf",
                        use_container_width=True
                    )
                except Exception as ex:
                    st.error(f"Erro ao gerar PDF: {ex}")



# ════════════════════════════════════════════════════════════
#  RODAPÉ
# ════════════════════════════════════════════════════════════
st.markdown("<br>", unsafe_allow_html=True)
st.markdown(
    f"<hr style='border:1px solid {BORDER}'>"
    f"<p style='text-align:center;color:{BORDER};font-size:.72rem;letter-spacing:2px'>"
    f"JBS SNIPER · FERRAMENTA EXCLUSIVA · {datetime.now().year}</p>",
    unsafe_allow_html=True)
